from tkinter import *
from tkinter import ttk
import mysql.connector
from tkinter import messagebox
from tkinter import filedialog
import openpyxl

class GPA:
    def __init__(self,root):
        self.root = root
        self.root.title('Student Record System')
        self.root.geometry('1600x800+0+0')
        
        
        self.Myanmar_Credit = IntVar()
        self.English_Credit= IntVar()
        self.Math_Credit= IntVar()
        self.Physics_Credit = IntVar()
        self.Chemistry_Credit = IntVar()
        self.Biology_Credit = IntVar()
        self.RollNo = StringVar()
        self.Name = StringVar()

        self.Myanmar_Grade = IntVar()
        self.English_Grade = IntVar()
        self.Math_Grade = IntVar()
        self.Physics_Grade = IntVar()
        self.Chemistry_Grade = IntVar()
        self.Biology_Grade = IntVar()
        
        self.Myanmar_GradeScore = IntVar()
        self.English_GradeScore = IntVar()
        self.Math_GradeScore = IntVar()
        self.Physics_GradeScore = IntVar()
        self.Chemistry_GradeScore = IntVar()
        self.Biology_GradeScore = IntVar()
        
        self.Myanmar_GradePoint = IntVar()
        self.English_GradePoint = IntVar()
        self.Math_GradePoint = IntVar()
        self.Physics_GradePoint = IntVar()
        self.Chemistry_GradePoint = IntVar()
        self.Biology_GradePoint = IntVar()
        
        self.Total_Credits=StringVar()
        self.Total_GradePoints=StringVar()
        self.gpa=StringVar()
        
        self.mydb = mysql.connector.connect(
            host = "localhost",
            user = "root",
            passwd = "root",
            database = "project"
        )

        self.mycursor = self.mydb.cursor()

        self.MainFrame = Frame(self.root, bg="white")
        self.MainFrame.pack(fill=BOTH, expand=1, padx=20, pady=20)

        # Title
        self.lblTitle = Label(self.MainFrame, font=('Arial', 20, 'bold'), text='GPA Sheet', padx=2, pady=2, bg="white")
        self.lblTitle.grid(row=0, column=0, columnspan=5, pady=(0, 20))

        ##########################Widget...........
        self.rollS = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Enter Roll No That You Want To Add Mark', padx=2, pady=2, bg="white")
        self.rollS.grid(row=1, column=0, sticky='w')
        self.txtSearchroll = Entry(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.RollNo, width=28)
        self.txtSearchroll.grid(row=1, column=1)

        self.lbName = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Name', padx=2, pady=2, bg="white")
        self.lbName.grid(row=2, column=0, sticky='w')
        self.txtName = Entry(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Name, width=28)
        self.txtName.grid(row=2, column=1)

        # Header row
        self.lbSubject = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Subject', padx=2, pady=2, bg="white")
        self.lbSubject.grid(row=3, column=0, sticky='w')
        
        self.lbCredit = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Credit Unit', padx=2, pady=2, bg="white")
        self.lbCredit.grid(row=3, column=1, sticky='w')
        
        self.lbGrade = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Grade', padx=2, pady=2, bg="white")
        self.lbGrade.grid(row=3, column=2, sticky='w')
        
        self.lbGradeScore = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Grade Score', padx=2, pady=2, bg="white")
        self.lbGradeScore.grid(row=3, column=3, sticky='w')
        
        self.lbGradePoint = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Grade Point', padx=2, pady=2, bg="white")
        self.lbGradePoint.grid(row=3, column=4, sticky='w')

        #Total credit point
        self.lbTotalCreditPoint = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Total Credit Point', padx=2, pady=2, bg="white")
        self.lbTotalCreditPoint.grid(row=10, column=0, sticky='w')
        self.txttotalCredit = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Total_Credits, width=28)
        self.txttotalCredit.grid(row=10, column=1, sticky='w')

        # Total grade point
        self.lbTotalGradePoint = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Total Grade Point', padx=2, pady=2, bg="white")
        self.lbTotalGradePoint.grid(row=10, column=3, sticky='w')
        self.txttotalgradePoint = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Total_GradePoints, width=28)
        self.txttotalgradePoint.grid(row=10, column=4, sticky='w')

        # GPA
        self.lbTotalgpa = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Overall GPA', padx=2, pady=2, bg="white")
        self.lbTotalgpa.grid(row=11, column=3, sticky='w')
        self.txttotalgpa = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.gpa, width=28)
        self.txttotalgpa.grid(row=11, column=4, sticky='w')


        # Myanmar
        self.lbMyanmar = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Myanmar', padx=2, pady=2, bg="white")
        self.lbMyanmar.grid(row=4, column=0, sticky='w')
        self.txtMyanmarCredit = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Myanmar_Credit, width=28)
        self.txtMyanmarCredit.grid(row=4, column=1 , sticky='w')
        self.txtMyanmarGrade = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Myanmar_Grade, width=28)
        self.txtMyanmarGrade.grid(row=4, column=2 , sticky='w')
        self.txtMyanmarGradeScore = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Myanmar_GradeScore, width=28)
        self.txtMyanmarGradeScore.grid(row=4, column=3 , sticky='w')
        self.txtMyanmarGradePoint = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Myanmar_GradePoint, width=28)
        self.txtMyanmarGradePoint.grid(row=4, column=4 , sticky='w')

        # English
        self.lbEnglish = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='English', padx=2, pady=2, bg="white")
        self.lbEnglish.grid(row=5, column=0, sticky='w')
        self.txtEnglishCredit = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.English_Credit, width=28)
        self.txtEnglishCredit.grid(row=5, column=1)
        self.txtEnglishGrade = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.English_Grade, width=28)
        self.txtEnglishGrade.grid(row=5, column=2)
        self.txtEnglishGradeScore = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.English_GradeScore, width=28)
        self.txtEnglishGradeScore.grid(row=5, column=3)
        self.txtEnglishGradePoint = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.English_GradePoint, width=28)
        self.txtEnglishGradePoint.grid(row=5, column=4)

        # Math
        self.lbMath = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Math', padx=2, pady=2, bg="white")
        self.lbMath.grid(row=6, column=0, sticky='w')
        self.txtMathCredit = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Math_Credit, width=28)
        self.txtMathCredit.grid(row=6, column=1)
        self.txtMathGrade = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Math_Grade, width=28)
        self.txtMathGrade.grid(row=6, column=2)
        self.txtMathGradeScore = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Math_GradeScore, width=28)
        self.txtMathGradeScore.grid(row=6, column=3)
        self.txtMathGradePoint = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Math_GradePoint, width=28)
        self.txtMathGradePoint.grid(row=6, column=4)

        # Physics
        self.lblPhysics = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Physics', padx=2, pady=2, bg="white")
        self.lblPhysics.grid(row=7, column=0, sticky='w')
        self.txtPhysicsCredit = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Physics_Credit, width=28)
        self.txtPhysicsCredit.grid(row=7, column=1)
        self.txtPhysicsGrade = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Physics_Grade, width=28)
        self.txtPhysicsGrade.grid(row=7, column=2)
        self.txtPhysicsGradeScore = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Physics_GradeScore, width=28)
        self.txtPhysicsGradeScore.grid(row=7, column=3)
        self.txtPhysicsGradePoint = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Physics_GradePoint, width=28)
        self.txtPhysicsGradePoint.grid(row=7, column=4)

        # Chemistry
        self.lblChemistry = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Chemistry', padx=2, pady=2, bg="white")
        self.lblChemistry.grid(row=8, column=0, sticky='w')
        self.txtChemistryCredit = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Chemistry_Credit, width=28)
        self.txtChemistryCredit.grid(row=8, column=1)
        self.txtChemistryGrade = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Chemistry_Grade, width=28)
        self.txtChemistryGrade.grid(row=8, column=2)
        self.txtChemistryGradeScore = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Chemistry_GradeScore, width=28)
        self.txtChemistryGradeScore.grid(row=8, column=3)
        self.txtChemistryGradePoint = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Chemistry_GradePoint, width=28)
        self.txtChemistryGradePoint.grid(row=8, column=4)

        # Biology
        self.lblBiology = Label(self.MainFrame, font=('Arial', 14, 'bold'), text='Biology', padx=2, pady=2, bg="white")
        self.lblBiology.grid(row=9, column=0, sticky='w')
        self.txtBiologyCredit = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Biology_Credit, width=28)
        self.txtBiologyCredit.grid(row=9, column=1)
        self.txtBiologyGrade = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Biology_Grade, width=28)
        self.txtBiologyGrade.grid(row=9, column=2)
        self.txtBiologyGradeScore = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Biology_GradeScore, width=28)
        self.txtBiologyGradeScore.grid(row=9, column=3)
        self.txtBiologyGradePoint = Label(self.MainFrame, font=('Arial', 14, 'bold'), textvariable=self.Biology_GradePoint, width=28)
        self.txtBiologyGradePoint.grid(row=9, column=4)

        ## Buttons
        self.btnSearch = Button(self.MainFrame, text='Calculate GPA', font=('Arial', 12, 'bold'), height=1, width=16, bd=2, padx=13, bg="#4CAF50", fg="white", command=self.Cal_GPA)
        self.btnSearch.grid(row=12, column=0, pady=(20, 5), padx=(0, 5))

        self.btnGPA = Button(self.MainFrame, text='Export GPA', font=('Arial', 12, 'bold'), height=1, width=16, bd=2, padx=13, bg="#2196F3", fg="white", command=self.Exp_GPA)
        self.btnGPA.grid(row=12, column=1, pady=(20, 5), padx=(5, 5))

        self.btnHome = Button(self.MainFrame, text='Home', font=('Arial', 12, 'bold'), height=1, width=16, bd=2, padx=13, bg="#FFC107", fg="black", command=self.Home)
        self.btnHome.grid(row=12, column=2, pady=(20, 5), padx=(5, 5))

        self.btnExit = Button(self.MainFrame, text='Exit', font=('Arial', 12, 'bold'), height=1, width=16, bd=2, padx=13, bg="#F44336", fg="white", command=self.Exit)
        self.btnExit.grid(row=12, column=3, pady=(20, 5), padx=(5, 0))

        # # Upload Photo Button
        # self.btnUploadPhoto = Button(self.MainFrame, text='Upload Photo', font=('Arial', 12, 'bold'), height=1, width=16, bd=2, padx=13, bg="#9C27B0", fg="white", command=self.upload_photo)
        # self.btnUploadPhoto.grid(row=13, column=0, pady=5, columnspan=4)


    def Home(self):
        from AddMark_Student_page import AddMark
        AddWindow = Toplevel(self.root)
        AddMark(AddWindow)

    def Exit(self):
        self.root.destroy()

    def Cal_GPA(self):
         # Get the roll number entered by the user for searching
        roll_to_search = self.RollNo.get()
        name_to_search=self.Name.get()
        # Write your SQL query to fetch student details based on the roll number
        query = "SELECT * FROM Student_Mark WHERE RollNo = %s and Name= %s"
        self.mycursor.execute(query, (roll_to_search,name_to_search))
        studentMark = self.mycursor.fetchone()

        # Populate the Entry fields with the fetched student details
        if studentMark:
            # Display the roll number
            roll=studentMark[0]
            name=studentMark[1]
            # Set the fetched student details to the respective variables or directly to Entry widgets
            myanmar=studentMark[2]
            english=studentMark[3]
            math=studentMark[4]
            physics=studentMark[5]
            chemistry=studentMark[6]
            biology=studentMark[7]

            if myanmar>=90:
                Myanmar_Credit=3
                Myanmar_GradeScore=4.0
                Myanmar_Grade='A+'
                Myanmar_GradePoint=Myanmar_Credit*Myanmar_GradeScore
            elif 89>=myanmar>=80:
                Myanmar_Credit=3
                Myanmar_GradeScore=4.0
                Myanmar_Grade='A'
                Myanmar_GradePoint=Myanmar_Credit*Myanmar_GradeScore
            elif 79>=myanmar>=75:
                Myanmar_Credit=3
                Myanmar_GradeScore=3.67
                Myanmar_Grade='A-'
                Myanmar_GradePoint=Myanmar_Credit*Myanmar_GradeScore
            elif 74>=myanmar>=70:
                Myanmar_Credit=3
                Myanmar_GradeScore=3.33
                Myanmar_Grade='B+'
                Myanmar_GradePoint=Myanmar_Credit*Myanmar_GradeScore
            elif 69>=myanmar>=65:
                Myanmar_Credit=3
                Myanmar_GradeScore=3.0
                Myanmar_Grade='B'
                Myanmar_GradePoint=Myanmar_Credit*Myanmar_GradeScore
            elif 64>=myanmar>=60:
                Myanmar_Credit=3
                Myanmar_GradeScore=2.67
                Myanmar_Grade='B-'
                Myanmar_GradePoint=Myanmar_Credit*Myanmar_GradeScore
            elif 59>=myanmar>=55:
                Myanmar_Credit=3
                Myanmar_GradeScore=2.33
                Myanmar_Grade='C+'
                Myanmar_GradePoint=Myanmar_Credit*Myanmar_GradeScore
            elif 54>=myanmar>=50:
                Myanmar_Credit=3
                Myanmar_GradeScore=2.0
                Myanmar_Grade='C'
                Myanmar_GradePoint=Myanmar_Credit*Myanmar_GradeScore
            elif 49>=myanmar:
                Myanmar_Credit=3
                Myanmar_GradeScore=1.67
                Myanmar_Grade='D'
                Myanmar_GradePoint=Myanmar_Credit*Myanmar_GradeScore

            if english>=90:
                English_Credit=3
                English_GradeScore=4.0
                English_Grade='A+'
                English_GradePoint=English_Credit*English_GradeScore
            elif 89>=english>=80:
                English_Credit=3
                English_GradeScore=4.0
                English_Grade='A'
                English_GradePoint=English_Credit*English_GradeScore
            elif 79>=english>=75:
                English_Credit=3
                English_GradeScore=3.67
                English_Grade='A-'
                English_GradePoint=English_Credit*English_GradeScore
            elif 74>=english>=70:
                English_Credit=3
                English_GradeScore=3.33
                English_Grade='B+'
                English_GradePoint=English_Credit*English_GradeScore
            elif 69>=english>=65:
                English_Credit=3
                English_GradeScore=3.0
                English_Grade='B'
                English_GradePoint=English_Credit*English_GradeScore
            elif 64>=english>=60:
                English_Credit=3
                English_GradeScore=2.67
                English_Grade='B-'
                English_GradePoint=English_Credit*English_GradeScore
            elif 59>=english>=55:
                English_Credit=3
                English_GradeScore=2.33
                English_Grade='C+'
                English_GradePoint=English_Credit*English_GradeScore
            elif 54>=english>=50:
                English_Credit=3
                English_GradeScore=2.0
                English_Grade='C'
                English_GradePoint=English_Credit*English_GradeScore
            elif 49>=english:
                English_Credit=3
                English_GradeScore=1.67
                English_Grade='D'
                English_GradePoint=English_Credit*English_GradeScore

            if math>=90:
                Math_Credit=3
                Math_GradeScore=4.0
                Math_Grade='A+'
                Math_GradePoint=Math_Credit*Math_GradeScore
            elif 89>=math>=80:
                Math_Credit=3
                Math_GradeScore=4.0
                Math_Grade='A'
                Math_GradePoint=Math_Credit*Math_GradeScore
            elif 79>=math>=75:
                Math_Credit=3
                Math_GradeScore=3.67
                Math_Grade='A-'
                Math_GradePoint=Math_Credit*Math_GradeScore
            elif 74>=math>=70:
                Math_Credit=3
                Math_GradeScore=3.33
                Math_Grade='B+'
                Math_GradePoint=Math_Credit*Math_GradeScore
            elif 69>=math>=65:
                Math_Credit=3
                Math_GradeScore=3.0
                Math_Grade='B'
                Math_GradePoint=Math_Credit*Math_GradeScore
            elif 64>=math>=60:
                Math_Credit=3
                Math_GradeScore=2.67
                Math_Grade='B-'
                Math_GradePoint=Math_Credit*Math_GradeScore
            elif 59>=math>=55:
                Math_Credit=3
                Math_GradeScore=2.33
                Math_Grade='C+'
                Math_GradePoint=Math_Credit*Math_GradeScore
            elif 54>=math>=50:
                Math_Credit=3
                Math_GradeScore=2.0
                Math_Grade='C'
                Math_GradePoint=Math_Credit*Math_GradeScore
            elif 49>=math:
                Math_Credit=3
                Math_GradeScore=1.67
                Math_Grade='D'
                Math_GradePoint=Math_Credit*Math_GradeScore

            if physics>=90:
                Physics_Credit=3
                Physics_GradeScore=4.0
                Physics_Grade='A+'
                Physics_GradePoint=Physics_Credit*Physics_GradeScore
            elif 89>=physics>=80:
                Physics_Credit=3
                Physics_GradeScore=4.0
                Physics_Grade='A'
                Physics_GradePoint=Physics_Credit*Physics_GradeScore
            elif 79>=physics>=75:
                Physics_Credit=3
                Physics_GradeScore=3.67
                Physics_Grade='A-'
                Physics_GradePoint=Physics_Credit*Physics_GradeScore
            elif 74>=physics>=70:
                Physics_Credit=3
                Physics_GradeScore=3.33
                Physics_Grade='B+'
                Physics_GradePoint=Physics_Credit*Physics_GradeScore
            elif 69>=physics>=65:
                Physics_Credit=3
                Physics_GradeScore=3.0
                Physics_Grade='B'
                Physics_GradePoint=Physics_Credit*Physics_GradeScore
            elif 64>=physics>=60:
                Physics_Credit=3
                Physics_GradeScore=2.67
                Physics_Grade='B-'
                Physics_GradePoint=Physics_Credit*Physics_GradeScore
            elif 59>=physics>=55:
                Physics_Credit=3
                Physics_GradeScore=2.33
                Physics_Grade='C+'
                Physics_GradePoint=Physics_Credit*Physics_GradeScore
            elif 54>=physics>=50:
                Physics_Credit=3
                Physics_GradeScore=2.0
                Physics_Grade='C'
                Physics_GradePoint=Physics_Credit*Physics_GradeScore
            elif 49>=physics:
                Physics_Credit=3
                Physics_GradeScore=1.67
                Physics_Grade='D'
                Physics_GradePoint=Physics_Credit*Physics_GradeScore

            if chemistry>=90:
                Chemistry_Credit=3
                Chemistry_GradeScore=4.0
                Chemistry_Grade='A+'
                Chemistry_GradePoint=Chemistry_Credit*Chemistry_GradeScore
            elif 89>=english>=80:
                Chemistry_Credit=3
                Chemistry_GradeScore=4.0
                Chemistry_Grade='A'
                Chemistry_GradePoint=Chemistry_Credit*Chemistry_GradeScore
            elif 79>=english>=75:
                Chemistry_Credit=3
                Chemistry_GradeScore=3.67
                Chemistry_Grade='A-'
                Chemistry_GradePoint=Chemistry_Credit*Chemistry_GradeScore
            elif 74>=english>=70:
                Chemistry_Credit=3
                Chemistry_GradeScore=3.33
                Chemistry_Grade='B+'
                Chemistry_GradePoint=Chemistry_Credit*Chemistry_GradeScore
            elif 69>=english>=65:
                Chemistry_Credit=3
                Chemistry_GradeScore=3.0
                Chemistry_Grade='B'
                Chemistry_GradePoint=Chemistry_Credit*Chemistry_GradeScore
            elif 64>=english>=60:
                Chemistry_Credit=3
                Chemistry_GradeScore=2.67
                Chemistry_Grade='B-'
                Chemistry_GradePoint=Chemistry_Credit*Chemistry_GradeScore
            elif 59>=english>=55:
                Chemistry_Credit=3
                Chemistry_GradeScore=2.33
                Chemistry_Grade='C+'
                Chemistry_GradePoint=Chemistry_Credit*Chemistry_GradeScore
            elif 54>=english>=50:
                Chemistry_Credit=3
                Chemistry_GradeScore=2.0
                Chemistry_Grade='C'
                Chemistry_GradePoint=Chemistry_Credit*Chemistry_GradeScore
            elif 49>=english:
                Chemistry_Credit=3
                Chemistry_GradeScore=1.67
                Chemistry_Grade='D'
                Chemistry_GradePoint=Chemistry_Credit*Chemistry_GradeScore

            if biology>=90:
                Biology_Credit=3
                Biology_GradeScore=4.0
                Biology_Grade='A+'
                Biology_GradePoint=Biology_Credit*Biology_GradeScore
            elif 89>=biology>=80:
                Biology_Credit=3
                Biology_GradeScore=4.0
                Biology_Grade='A'
                Biology_GradePoint=Biology_Credit*Biology_GradeScore
            elif 79>=biology>=75:
                Biology_Credit=3
                Biology_GradeScore=3.67
                Biology_Grade='A-'
                Biology_GradePoint=Biology_Credit*Biology_GradeScore
            elif 74>=biology>=70:
                Biology_Credit=3
                Biology_GradeScore=3.33
                Biology_Grade='B+'
                Biology_GradePoint=Biology_Credit*Biology_GradeScore
            elif 69>=biology>=65:
                Biology_Credit=3
                Biology_GradeScore=3.0
                Biology_Grade='B'
                Biology_GradePoint=Biology_Credit*Biology_GradeScore
            elif 64>=biology>=60:
                Biology_Credit=3
                Biology_GradeScore=2.67
                Biology_Grade='B-'
                Biology_GradePoint=Biology_Credit*Biology_GradeScore
            elif 59>=biology>=55:
                Biology_Credit=3
                Biology_GradeScore=2.33
                Biology_Grade='C+'
                Biology_GradePoint=Biology_Credit*Biology_GradeScore
            elif 54>=biology>=50:
                Biology_Credit=3
                Biology_GradeScore=2.0
                Biology_Grade='C'
                Biology_GradePoint=Biology_Credit*Biology_GradeScore
            elif 49>=biology:
                Biology_Credit=3
                Biology_GradeScore=1.67
                Biology_Grade='D'
                Biology_GradePoint=Biology_Credit*Biology_GradeScore

            self.Myanmar_Credit.set(Myanmar_Credit)
            self.Myanmar_Grade.set(Myanmar_Grade)
            self.Myanmar_GradeScore.set(Myanmar_GradeScore)
            self.Myanmar_GradePoint.set(Myanmar_GradePoint)
            
            self.English_Credit.set(English_Credit)
            self.English_Grade.set(English_Grade)
            self.English_GradeScore.set(English_GradeScore)
            self.English_GradePoint.set(English_GradePoint)
            
            self.Math_Credit.set(Math_Credit)
            self.Math_Grade.set(Math_Grade)
            self.Math_GradeScore.set(Math_GradeScore)
            self.Math_GradePoint.set(Math_GradePoint)
            
            self.Physics_Credit.set(Physics_Credit)
            self.Physics_Grade.set(Physics_Grade)
            self.Physics_GradeScore.set(Physics_GradeScore)
            self.Physics_GradePoint.set(Physics_GradePoint)
            
            self.Chemistry_Credit.set(Chemistry_Credit)
            self.Chemistry_Grade.set(Chemistry_Grade)
            self.Chemistry_GradeScore.set(Chemistry_GradeScore)
            self.Chemistry_GradePoint.set(Chemistry_GradePoint)
            
            self.Biology_Credit.set(Biology_Credit)
            self.Biology_Grade.set(Biology_Grade)
            self.Biology_GradeScore.set(Biology_GradeScore)
            self.Biology_GradePoint.set(Biology_GradePoint)
            
            # Calculate total credits and total grade points
            Total_Credits = (Myanmar_Credit + English_Credit + Math_Credit + Physics_Credit + Chemistry_Credit + Biology_Credit)
            Total_GradePoints = round((Myanmar_GradePoint + English_GradePoint + Math_GradePoint + Physics_GradePoint + Chemistry_GradePoint + Biology_GradePoint),2)
            
            self.Total_Credits.set(Total_Credits)
            self.Total_GradePoints.set(Total_GradePoints)
            # Calculate GPA
            gpa = round(Total_GradePoints / Total_Credits,3)
            self.gpa.set(gpa)
            
            # Display GPA in a message box
            messagebox.showinfo("GPA Calculation", f"The calculated GPA is: {gpa:.2f}")
        else:
            messagebox.showerror("Error", "Student record not found.")

    def Exp_GPA(self):
        # Create a new Excel workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GPA"

        # Define the headers for the Excel file
        ws['A1'] = "Roll No"
        ws['B1'] = "Name"
        ws['C1'] = "Subjects"
        ws['D1'] = "Credit"
        ws['E1'] = "Grade"
        ws['F1'] = "Grade Score"
        ws['G1'] = "Grade Point"

        # Save the workbook to a file named 'GPA.xlsx'
        wb.save(filename='GPA.xlsx')

        # Display a message box to indicate success
        root = Tk()
        root.withdraw()  # Hide the root window
        messagebox.showinfo("Exported", "GPA has been exported successfully!")
        root.destroy()  # Destroy the root window after the message box is closed
   
            
        

        
            
    def upload_photo(self):
        filename = filedialog.askopenfilename()
        print("Selected File:", filename)

    def Exp_GPA(self):
        # Implement GPA export logic here if required
        print('Hello')

if __name__ == '__main__':
    root = Tk()
    application = GPA(root)
    root.mainloop()
